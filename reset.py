import subprocess
import paramiko
import time
import re
from datetime import datetime as dt
import csv
import sys
from tkinter import *
from collections import OrderedDict
from broadcast_helper import broadcast_main
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import argparse


device_status = {}

def message_format(device_id, key, value):
    if device_id in device_status.keys():
        device_status[device_id][key] = value
    else:
        device_status[device_id] = OrderedDict()
        device_status[device_id][key] = value
    mas_len = 30
    space = " "*(mas_len-len(key)-4)
    return f"{key}{space}:{value}"

class Config:
    def __init__(self, product_line = "BGR",device_id = ""):
        self.device_id = device_id
        self.product_line = product_line
        self.username = "root"
        self.password = "EKM2800123Netra"
        self.bagheera_override = "/home/ubuntu/config/bagheera_override.ini"
        self.sam_config = "/home/ubuntu/.nddevice/latest/sam_config.ini"
        if product_line in "KRT":
            self.password = "EKM2020123Krait"
            self.bagheera_override = "/data/nd_files/config/bagheera_override.ini"

class SSHSession:
    def __init__(self, host, port=22, username="", password=""):
        self.host = host
        self.port = port
        self.username = username
        self.password = password
        self.client = paramiko.SSHClient()
        self.client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        print(f"Connecting to {self.host}:{self.port} as {self.username}")
        self.client.connect(self.host, self.port, self.username, self.password)

    def execute_command(self, command, timeout=20):
        stdin, stdout, stderr = self.client.exec_command(command,timeout=timeout)
        return stdout.read().decode("utf-8"), stderr.read().decode("utf-8")

    def upload_file(self, local_file_path, remote_file_path):
        # ftp = self.client.open_sftp()
        # ftp.put(local_file_path, remote_file_path)
        command = f"sshpass -p {self.password} scp {local_file_path} {self.username}@{self.host}:{remote_file_path}"
        p = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = p.communicate()
        if "Host key verification failed." in stderr.decode():
            print("Host key verification failed.")
        self.execute_command("sync")
        time.sleep(0.5)
        # ftp.close()

    def download_file(self, remote_file_path, local_file_path):
        ftp = self.client.open_sftp()
        ftp.get(remote_file_path, local_file_path)
        ftp.close()

    def __del__(self):
        self.client.close()

class Setup:
    def __init__(self, ssh: SSHSession, config:Config):
        self.ssh: SSHSession = ssh
        self.config: Config  = config

    def setup_sam_config(self):
        self.ssh.upload_file("configs/sam_config.ini" ,self.config.sam_config)
        output, err = self.ssh.execute_command(f"md5sum {self.config.sam_config}")
        remote_output_md5sum = re.search(r"([a-f0-9]{32})", output)
        local_output = subprocess.run(["md5sum", "configs/sam_config.ini"], stdout=subprocess.PIPE ).stdout.decode("utf-8")
        local_output_md5sum = re.search(r"([a-f0-9]{32})", local_output)
        if (remote_output_md5sum.group(1) == local_output_md5sum.group(1)):
            print(message_format(self.config.device_id,"sam_config.ini", "PASS"))
        else:
            print(err)
            print(message_format(self.config.device_id,"sam_config.ini", "FAIL"))

    def setup_conn_mgr_config(self):
        self.ssh.upload_file("configs/conn_mgr_config.txt" ,"/home/ubuntu/config/conn_mgr_config.txt")
        output, err = self.ssh.execute_command("cat /home/ubuntu/config/conn_mgr_config.txt")
        if ("SPRINT" in output) and ("internet.curiosity.sprint.com" in output):
            print(message_format(self.config.device_id,"conn_mgr_config.txt", "PASS"))
        else:
            print(message_format(self.config.device_id,"conn_mgr_config.txt", "FAIL"))

    def setup_bagheera_override(self):
        self.ssh.upload_file("configs/bagheera_override.ini", self.config.bagheera_override)
        output, err = self.ssh.execute_command(f"md5sum {self.config.bagheera_override}")
        remote_output_md5sum = re.search(r"([a-f0-9]{32})", output)
        local_output = subprocess.run(["md5sum", "configs/bagheera_override.ini"], stdout=subprocess.PIPE ).stdout.decode("utf-8")
        local_outpu_md5sum = re.search(r"([a-f0-9]{32})", local_output)
        try:
            if remote_output_md5sum.group(1) == local_outpu_md5sum.group(1):
                print(message_format(self.config.device_id,"bagheera_override.ini", "PASS"))
            else:
                print(message_format(self.config.device_id,"bagheera_override.ini", "FAIL"))
        except Exception as e:
            print(e)
            print(message_format(self.config.device_id,"bagheera_override.ini", "FAIL"))

    def check_event_logs_vod_obs(self):
        if self.config.product_line == "KRT":
            event_command = "sqlite3 /home/ubuntu/.nddevice/db/uploader.db 'select * from uploader' |wc -l"
        else:
            event_command = "sqlite3 /home/ubuntu/.nddevice/uploader.db 'select * from uploader' |wc -l"
        output, err = self.ssh.execute_command(event_command)
        print(message_format(self.config.device_id,"Event Upload remaining", output.strip()))

        if self.config.product_line == "KRT":
            vod_command = "sqlite3 /home/ubuntu/.nddevice/db/uploader.db 'select * from uploader_vod' |wc -l"
        else:
            vod_command = "sqlite3 /home/ubuntu/.nddevice/uploader.db 'select * from uploader_vod' |wc -l"
        output, err = self.ssh.execute_command(vod_command)
        print(message_format(self.config.device_id,"VOD upload remaining", output.strip()))

        critical_archive_log_command = "ls /home/ubuntu/.nddevice/log/archive/critical/ | wc -l"
        output, err = self.ssh.execute_command(critical_archive_log_command)
        print(message_format(self.config.device_id,"Critical Archive Log", output.strip()))

        critical_vod_command = "ls /home/ubuntu/.nddevice/observations/ | wc -l"
        output, err = self.ssh.execute_command(critical_vod_command)
        print(message_format(self.config.device_id,"Obs upload remaining", output.strip()))

    def check_nd_output_nd_input(self):
        if self.config.product_line == "KRT":
            output,err = self.ssh.execute_command("echo 'EKM2020123Krait' | sudo rm -rf /home/iriscli/ND_OUTPUT/*")
            output_in,err = self.ssh.execute_command("echo 'EKM2020123Krait' | sudo rm -rf /home/iriscli/ND_INPUT/*")
        else:
            output,err = self.ssh.execute_command("echo 'EKM2800123Netra' | sudo rm -rf /home/iriscli/ND_OUTPUT/*")
            output_in,err = self.ssh.execute_command("echo 'EKM2800123Netra' | sudo rm -rf /home/iriscli/ND_INPUT/*")

        nd_out_command = "ls /home/iriscli/ND_OUTPUT/ | wc -l"
        nd_in_command = "ls /home/iriscli/ND_INPUT/ | wc -l"
        output_out, err = self.ssh.execute_command(nd_out_command)
        output_in, err = self.ssh.execute_command(nd_in_command)
        print(message_format(self.config.device_id,"ND_output", output_out.strip()))
        print(message_format(self.config.device_id,"ND_input", output_in.strip()))

    def setup_certificates(self):
        # overall_status = False
        # certificate_checks = {
        #     "private.pem.key": False,
        #     "certificate.pem.crt": False,
        #     "ed25519key.pem": False,
        #     "pub-ed25519.pem": False
        # }
        max_len = 30

        output, err = self.ssh.execute_command("cp /home/ubuntu/.nddevice/certificate/backup/cacert.pem /home/ubuntu/.nddevice/certificate/cacert.pem")
        output, err = self.ssh.execute_command("rm /home/ubuntu/.nddevice/certificate/temp.txt")
        output, err = self.ssh.execute_command("rm /home/ubuntu/.nddevice/certificate/private.pem.key")
        output, err = self.ssh.execute_command("rm /home/ubuntu/.nddevice/certificate/certificate.pem.crt")
        output, err = self.ssh.execute_command("rm /home/ubuntu/.nddevice/certificate/ed25519key.pem")
        output, err = self.ssh.execute_command("rm /home/ubuntu/.nddevice/certificate/pub-ed25519.pem")
        output, err = self.ssh.execute_command("sync")
        time.sleep(0.5)
        output, err = self.ssh.execute_command("systemctl restart awsiot")
        time.sleep(20)
        output,err =  self.ssh.execute_command("ls /home/ubuntu/.nddevice/certificate/private.pem.key")
        if err != "":
            print(message_format(self.config.device_id,"private.pem.key", "FAIL"))
        else:
            print(message_format(self.config.device_id,"private.pem.key", "PASS"))
        output,err =  self.ssh.execute_command("ls /home/ubuntu/.nddevice/certificate/certificate.pem.crt")
        if err != "":
            print(message_format(self.config.device_id,"certificate.pem.crt", "FAIL"))
        else:
            print(message_format(self.config.device_id,"certificate.pem.crt", "PASS"))
        output,err =  self.ssh.execute_command("ls /home/ubuntu/.nddevice/certificate/ed25519key.pem")
        if err != "":
            print(message_format(self.config.device_id,"ed25519key", "FAIL"))
        else:
            print(message_format(self.config.device_id,"ed25519key", "PASS"))
        output,err =  self.ssh.execute_command("ls /home/ubuntu/.nddevice/certificate/pub-ed25519.pem")
        if err != "":
            print(message_format(self.config.device_id,"pub-ed25519.pem", "FAIL"))
        else:
            print(message_format(self.config.device_id,"pub-ed25519.pem", "PASS"))


    def setup_services(self):
        active_service_list =  [
            "wifi_mgr",
            "uploader",
            "time_sync",
            "svc",
            "speed",
            "service_mon",
            "scheduler_manager",
            "power_monitor",
            "nd_dta",
            "nd_bt",
            "installer_app",
            "HealthStatsManager",
            "diagnostic",
            "conn_mgr",
            "circular_buffer",
            "cam_rec",
            "bagheera",
            "unifiedAnalyticsClient",
            "outwardAnalyticsClient",
            "audioPlayback",
            "analytics*",
            "awsiot",
            "obd",
        ]
        if self.config.product_line == "KRT":
            active_service_list.remove("cam_rec")
            message_format(self.config.device_id, "cam_rec", "N/A")
        deactive_service_list = [
            "nd_sam",
            "ext_cam",
        ]
        max_len = 30
        for service in active_service_list:
            output, err = self.ssh.execute_command(f"systemctl status {service}")
            if "active (running)" in output:
                print(message_format(self.config.device_id, service, "PASS"))
            else:
                print(message_format(self.config.device_id, service, "FAIL"))
        for service in deactive_service_list:
            output, err = self.ssh.execute_command(f"systemctl status {service}")
            if "inactive (dead)" in output:
                print(message_format(self.config.device_id, service, "PASS"))
            else:
                print(message_format(self.config.device_id, service, "FAIL"))

    def device_date_check(self):
        output, err = self.ssh.execute_command("date")
        print(message_format(self.config.device_id, "Device Date", output.strip()))
        p = subprocess.Popen("date -u", shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        stdout, stderr = p.communicate()

        print(message_format(self.config.device_id, "UTC Date", stdout.decode().strip()))
        # device_date = dt.strptime(output.strip(), "%a %b %d %H:%M:%S %Z %Y")
        # # utc_date = dt.strptime(stdout.decode().strip(), "%A %d %B %Y %I:%M:%S %p %Z")
        # utc_date = dt.strptime(stdout.decode().strip(), "%A %d %B %Y %I:%M:%S %p %Z")

        # if device_date == utc_date:
        #     print(message_format(self.config.device_id, "Date Check", "PASS"))
        # else:
        #     if(self.config.product_line == "KRT"):
        #         utc_date_str = utc_date.strftime("%d %b %Y %H:%M:%S").upper()
        #         _,err = self.ssh.execute_command(f"date -s '{utc_date_str}';hwclock -w;hwclock")
        #     else:
        #         utc_date_str = utc_date.strftime("%m/%d/%Y %H:%M:%S")
        #         _,err = self.ssh.execute_command(f"sudo hwclock --set --date \"{utc_date_str}\" -f /dev/rtc;sudo hwclock --hctosys -f /dev/rtc;sudo hwclock -r")
            
        #     if err != "":
        #         print(message_format(self.config.device_id, "Date Check", "FAIL"))
        #     else:
        #         print(message_format(self.config.device_id, "Date Check", "Synced with UTC"))

    def get_device_version(self):
        output,err = self.ssh.execute_command("cat /home/ubuntu/.nddevice/nddevice.ini | grep nddevice | head -n 1 | awk -F'= ' {'print $2'}")
        print(message_format(self.config.device_id, "Device Version", output.strip()))
    
    def get_lumia_id(self):
        if self.config.product_line == "KRT":
            output,err = self.ssh.execute_command("echo 'EKM2020123Krait' | lte_gps_test 'ATi' | grep 'Model' | awk -F': ' {'print $2'}")
        else:
            output,err = self.ssh.execute_command("echo 'EKM2800123Netra' | sudo lte_gps_test 'ATi' | grep 'Model' | awk -F': ' {'print $2'}")
        print(message_format(self.config.device_id, "Lumia ID", output.strip()))

    def reboot(self):
        try:
            if self.config.product_line == "KRT":
                self.ssh.execute_command("echo 'msp qcs_reboot' | cli_mgr", timeout=10)
            elif self.config.product_line == "BGR":
                self.ssh.execute_command("reboot", timeout=10)
        except:
            pass

def reset_main(csv_file):
    rPi_set = set()

    with open(csv_file) as file:
        device_data = csv.DictReader(file)
        for device in device_data:
            device_id = str(device.get("device_id"))
            ip_address = str(device.get("ip_address"))
            rPi_set.add(str(device.get("username")))
            product_line = "BGR"
            if device_id.startswith("264"):
                product_line = "KRT"
            if device_id.startswith("66"):
                product_line = "KRT"
            config = Config(product_line,device_id)
            username = config.username
            password = config.password
            print(f"Device ID: {device_id}, IP Address: {ip_address}, Product Line: {product_line}") 
            print("--------------------------------------------------------------------")
            ssh = SSHSession(ip_address, 22, username, password)
            setup = Setup(ssh, config)

            setup.get_lumia_id()
            setup.get_device_version()
            setup.device_date_check()
            setup.check_event_logs_vod_obs()
            setup.check_nd_output_nd_input()
            setup.setup_sam_config()
            setup.setup_conn_mgr_config()
            setup.setup_bagheera_override()
            # setup.setup_certificates()
            setup.setup_services()
            setup.reboot()
            print("====================================================================")

def create_label(root, text, width, label_type="default"):
    if label_type == "header":
        return Label(root, text=text, relief=RIDGE, width=width, bg="lightgrey", font=("Helvetica", 12, "bold"), height=2, padx=2, pady=5)
    elif label_type == "pass":
        return Label(root, text=text, relief=RIDGE, width=width, bg="lightgreen", font=("Helvetica", 12),height=1, padx=2, pady=2)
    elif label_type == "fail":
        return Label(root, text=text, relief=RIDGE, width=width, bg="lightcoral", font=("Helvetica", 12),height=1, padx=2, pady=2)
    elif label_type == "leftHeader":
        return Label(root, text=text, relief=RIDGE, width=width, bg="white", font=("Helvetica", 12, "bold"), height=1, padx=2, pady=2)
    else:
        return Label(root, text=text, relief=RIDGE, width=width, bg="lightblue", font=("Helvetica", 12),height=1, padx=2, pady=2)

def make_table(root, data):
    total_cols = len(data) + 1
    total_rows = len(next(iter(data.values()))) + 1
    row_names = list(next(iter(data.values())).keys())

    # Calculate the maximum width required for each column
    col_widths = [len("Device ID")]
    for device_id in data.keys():
        col_widths.append(len(device_id))
    for row_name in row_names:
        col_widths[0] = max(col_widths[0], len(row_name))
        for device_id in data.keys():
            value = data[device_id][row_name]
            col_widths[list(data.keys()).index(device_id) + 1] = max(col_widths[list(data.keys()).index(device_id) + 1], len(value))

    # Create a canvas and a frame inside it
    canvas = Canvas(root)
    frame = Frame(canvas)
    h_scrollbar = Scrollbar(root, orient=HORIZONTAL, command=canvas.xview)
    v_scrollbar = Scrollbar(root, orient=VERTICAL, command=canvas.yview)
    canvas.configure(xscrollcommand=h_scrollbar.set, yscrollcommand=v_scrollbar.set)

    h_scrollbar.pack(side=BOTTOM, fill=X)
    v_scrollbar.pack(side=RIGHT, fill=Y)
    canvas.pack(side=LEFT, fill=BOTH, expand=True)
    canvas.create_window((0, 0), window=frame, anchor="nw")

    # Update the scroll region of the canvas
    frame.bind("<Configure>", lambda event, canvas=canvas: canvas.configure(scrollregion=canvas.bbox("all")))

    for i in range(total_rows):
        if i == 0:
            for j in range(total_cols):
                if j == 0:
                    label = create_label(frame, "Device ID", col_widths[j], label_type="header")
                else:
                    label = create_label(frame, list(data.keys())[j-1], col_widths[j], label_type="header")
                label.grid(row=i, column=j)
        else:
            for j in range(total_cols):
                if j == 0:
                    label = create_label(frame, row_names[i-1], col_widths[j], label_type="leftHeader")
                else:
                    device_id = list(data.keys())[j-1]
                    value = data[device_id][row_names[i-1]]
                    label_type = "pass" if value == "PASS" else "fail" if value == "FAIL" else "default"
                    label = create_label(frame, value, col_widths[j], label_type=label_type)
                label.grid(row=i, column=j)

def save_table_to_excel(data, file_path):
    wb = Workbook()
    ws = wb.active

    # Define styles
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    pass_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    fail_fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
    left_header_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    default_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    header_font = Font(name="Helvetica", size=12, bold=True)
    default_font = Font(name="Helvetica", size=12)
    alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    headers = ["Device ID"] + list(data.keys())
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment

    # Write data
    row_names = list(next(iter(data.values())).keys())
    for row_num, row_name in enumerate(row_names, 2):
        cell = ws.cell(row=row_num, column=1, value=row_name)
        cell.fill = left_header_fill
        cell.font = header_font
        cell.alignment = alignment

        for col_num, device_id in enumerate(data.keys(), 2):
            value = data[device_id][row_name]
            cell = ws.cell(row=row_num, column=col_num, value=value)
            if value == "PASS":
                cell.fill = pass_fill
            elif value == "FAIL":
                cell.fill = fail_fill
            else:
                cell.fill = default_fill
            cell.font = default_font
            cell.alignment = alignment

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Device reset script")
    parser.add_argument("-d", "--devices", help="Comma-separated list of devices")
    parser.add_argument("-c", "--csv", help="CSV file with device information")
    parser.add_argument("-s", "--save", action="store_true", help="Save output to Excel")
    parser.add_argument("--no-tk", action="store_true", help="Disable Tkinter table display")

    args = parser.parse_args()

    print("Arguments parsed:", args)

    if args.devices:
        print("Devices argument provided")
        device_list = args.devices.split(",")
        broadcast_main(device_list)
        print("\n\n\n")
        reset_main("device.csv")
    elif args.csv:
        print("CSV argument provided")
        reset_main(args.csv)
    else:
        print("Please provide the device list with -d or the device csv with -c")

    if args.save:
        print("Saving to Excel")
        save_table_to_excel(device_status, "device_status.xlsx")

    if not args.no_tk:
        print("Displaying Tkinter table")
        root = Tk()
        make_table(root, device_status)
        root.title("Device Status Table")
        root.geometry("1920x1080")
        root.mainloop()
